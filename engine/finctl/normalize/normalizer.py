"""Normalize arbitrary input into the canonical schema.

BEHAVIOR.md, stage `normalize`:
  Promises  — canonical columns, integer paise, UTC. The ONLY place rupee strings parse.
  Refuses   — to guess a column mapping. Unrecognised or ambiguous columns raise.
  Bad input — empty file returns an empty frame WITH the schema; "nothing to reconcile"
              is a valid answer that must survive to the verdict stage.

The refusal to guess is the load-bearing part. A tool that silently maps a column it is
unsure about produces a confident, wrong reconciliation — and a merchant has no way to
tell that from a correct one. Raising costs seconds; guessing costs trust.
"""

from __future__ import annotations

import csv
import json
import warnings
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from finctl.money import MoneyError, parse_money
from finctl.schema import (
    BANK_ALIASES,
    BANK_REQUIRED,
    LEDGER_ALIASES,
    LEDGER_REQUIRED,
    normalise_key,
)


class NormalizationError(ValueError):
    """Raised when input cannot be mapped or parsed.

    Always names the offending column or value. An error that says only "bad input"
    forces a hunt through a 50,000-row file.
    """


class UnmappedColumnsError(NormalizationError):
    """A required column could not be mapped, carrying what a human needs to fix it.

    Subclasses NormalizationError so every existing `except NormalizationError` and
    every test matching on the message keeps working — the message is unchanged. What
    is added is the same information as STRUCTURED DATA, so a UI can render a picker
    instead of asking a merchant to read a paragraph and rename a column by hand.

    This is not a weakening of the refusal to guess. The engine still refuses. It now
    hands over the evidence for a HUMAN to decide, once, which is what Cointab and
    Hyperswitch both do at onboarding (docs/PRIOR-ART.md). See ADR-045.
    """

    def __init__(
        self,
        message: str,
        *,
        source: str,
        unmapped: list[str],
        headers: list[str],
        resolved: dict[str, str],
        aliases: dict[str, tuple[str, ...]],
    ) -> None:
        super().__init__(message)
        self.source = source
        self.unmapped = unmapped
        self.headers = headers
        self.resolved = resolved
        self.aliases = aliases

    def as_dict(self) -> dict[str, Any]:
        """What a mapping picker needs: what is missing, and what is available.

        `candidates` deliberately lists EVERY unclaimed header rather than a ranked
        guess. Ordering them by similarity would put a suggestion in front of a human
        who is being asked precisely because the engine cannot tell — and a plausible
        wrong suggestion accepted without thought is worse than no suggestion.
        """
        claimed = set(self.resolved.values())
        available = [h for h in self.headers if h not in claimed]
        return {
            "error": "unmapped_columns",
            "source": self.source,
            "message": str(self),
            "unmapped": [
                {
                    "canonical": name,
                    "accepted_spellings": list(self.aliases.get(name, ())),
                    "candidates": available,
                }
                for name in self.unmapped
            ],
            "already_mapped": dict(sorted(self.resolved.items())),
            "headers": list(self.headers),
        }


@dataclass(frozen=True)
class ColumnMapping:
    """The resolved input-column -> canonical-column map, kept for the audit trail.

    Recorded rather than discarded because "which column did you read as the amount?"
    is a question a merchant may reasonably ask when a number looks wrong.
    """

    resolved: dict[str, str]           # canonical -> actual input column name
    unmapped: tuple[str, ...]          # input columns we did not use
    # Fields a HUMAN mapped explicitly rather than the alias table resolving. Recorded
    # so the audit trail distinguishes "we recognised this column" from "someone told
    # us what this column was" — a different kind of claim, and worth being able to
    # tell apart when a number is disputed. See ADR-045.
    overridden: tuple[str, ...] = ()

    def describe(self) -> str:
        pairs = ", ".join(f"{v!r}->{k}" for k, v in sorted(self.resolved.items()))
        extra = f"; ignored {list(self.unmapped)}" if self.unmapped else ""
        chosen = f"; mapped by hand: {list(self.overridden)}" if self.overridden else ""
        return pairs + chosen + extra


def resolve_columns(
    headers: list[str],
    aliases: dict[str, tuple[str, ...]],
    required: tuple[str, ...],
    source_name: str,
    overrides: dict[str, str] | None = None,
) -> ColumnMapping:
    """Map input headers onto canonical names, or raise explaining why not.

    Never positional. A reordered file is fine; an unrecognisable one is an error.

    `overrides` is a canonical -> input-column map supplied by a HUMAN who was shown the
    unmapped columns and chose. It is applied BEFORE the alias table, and it wins: a
    person who has looked at their own export knows more about it than our alias list
    does. Every override is recorded in the ColumnMapping and reaches the audit trail,
    so "which column did you read as the amount?" stays answerable — and the answer
    names who decided. See ADR-045.
    """
    overrides = overrides or {}
    folded = {normalise_key(h): h for h in headers}
    if len(folded) != len(headers):
        seen: dict[str, list[str]] = {}
        for h in headers:
            seen.setdefault(normalise_key(h), []).append(h)
        dupes = {k: v for k, v in seen.items() if len(v) > 1}
        raise NormalizationError(
            f"{source_name}: columns collide after folding: {dupes}. "
            "Rename them so each maps to a distinct field."
        )

    resolved: dict[str, str] = {}
    claimed: set[str] = set()

    # Human choices first. An override naming a column that is not in the file is a
    # mistake worth refusing loudly: silently ignoring it would fall through to the
    # alias table and produce a mapping the person did not ask for.
    for canonical, input_col in overrides.items():
        if canonical not in aliases:
            raise NormalizationError(
                f"{source_name}: cannot map unknown field {canonical!r}. "
                f"Known fields: {sorted(aliases)}."
            )
        if input_col not in headers:
            raise NormalizationError(
                f"{source_name}: mapping for {canonical!r} names column {input_col!r}, "
                f"which is not in the file. Headers: {headers}."
            )
        if input_col in claimed:
            raise NormalizationError(
                f"{source_name}: column {input_col!r} is mapped to more than one field. "
                "Each input column may be used once."
            )
        resolved[canonical] = input_col
        claimed.add(input_col)

    human_mapped = set(resolved)

    for canonical, candidates in aliases.items():
        if canonical in resolved:
            continue   # a human already decided this one
        # dict.fromkeys de-duplicates while preserving order. Several aliases can fold
        # to the same key ("order_id" and "orderid" both fold to "orderid"), so the same
        # input column may be hit more than once - that is one candidate, not an
        # ambiguity. Only DISTINCT input columns constitute a genuine ambiguity.
        hits = list(
            dict.fromkeys(
                folded[normalise_key(c)] for c in candidates if normalise_key(c) in folded
            )
        )
        hits = [h for h in hits if h not in claimed]
        if not hits:
            continue
        if len(hits) > 1:
            # Ambiguity is reported, never resolved by preference order. Picking the
            # "best" candidate is exactly the silent guess this stage exists to refuse.
            raise NormalizationError(
                f"{source_name}: ambiguous mapping for {canonical!r} — "
                f"input has {sorted(hits)}, and more than one could be it. "
                "Rename or remove the extras so the intent is explicit."
            )
        resolved[canonical] = hits[0]
        claimed.add(hits[0])

    missing = [c for c in required if c not in resolved]
    if missing:
        raise UnmappedColumnsError(
            f"{source_name}: could not map required column(s) {missing}. "
            f"Input headers: {headers}. "
            f"Accepted spellings: "
            + "; ".join(f"{m}: {list(aliases[m])}" for m in missing)
            + ". Refusing to guess — see docs/BEHAVIOR.md, stage `normalize`.",
            source=source_name,
            unmapped=missing,
            headers=list(headers),
            resolved=dict(resolved),
            aliases=aliases,
        )

    return ColumnMapping(
        resolved=resolved,
        unmapped=tuple(h for h in headers if h not in claimed),
        overridden=tuple(sorted(human_mapped)),
    )


# Excel's day-zero. Serial 1 is 1900-01-01, but Excel wrongly treats 1900 as a leap
# year, so the epoch that makes modern dates come out right is 1899-12-30.
EXCEL_EPOCH = datetime(1899, 12, 30, tzinfo=UTC)

# The window in which a bare number is read as an Excel serial rather than epoch
# seconds. 20000 is 1954-10-03 and 80000 is 2119-01-25 — wider than any settlement
# file will contain, and four orders of magnitude below the epoch-seconds range for
# the same dates (2020-01-01 is serial 43831 but epoch 1577836800). The two
# interpretations are therefore separated by a gap of ~10^4, not adjacent ranges
# needing a judgement call. See ADR-037.
EXCEL_SERIAL_MIN = 20_000
EXCEL_SERIAL_MAX = 80_000


def _looks_numeric(text: str) -> bool:
    """True for a bare decimal number, e.g. "44658" or "44658.44689814815".

    Deliberately narrower than float(): it rejects "1e9", "inf", "nan" and signed
    values, none of which are dates, so they fall through to the string formats and
    ultimately to a raised error rather than being coerced.
    """
    if not text:
        return False
    whole, _, frac = text.partition(".")
    return whole.isdigit() and (frac == "" or frac.isdigit())


def _from_excel_serial(serial: float) -> datetime:
    """Convert an Excel/OOXML serial date to UTC.

    The fractional part is the time of day: 44658.44689814815 is 2022-04-07 10:43:32.
    """
    return EXCEL_EPOCH + timedelta(days=serial)


def _parse_timestamp(value: Any, source_name: str, row_num: int) -> datetime | None:
    """Parse a timestamp to UTC. Accepts Excel serials, epoch seconds, ISO, DD/MM/YYYY.

    The Excel-serial branch exists because Razorpay's own dashboard exports carry them.
    In `sample-settlements-recon-report.xlsx` the SAME column holds both
    `44658.44689814815` and `29/06/2022 07:34:39` — a spreadsheet writes whichever the
    cell format dictates, so a real file mixes the two. Reading 44658 as epoch seconds
    yields 1970-01-01, which is not an error a merchant would ever see raised: it is a
    plausible-looking date that quietly makes every settlement look years late. See
    ADR-037.
    """
    if value in (None, ""):
        return None

    # A real datetime can arrive already parsed (openpyxl hands back datetimes for
    # date-formatted cells). Nothing to do but normalise the timezone.
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=UTC)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC)

    text = str(value).strip()

    # Excel serial, integer or fractional. Checked BEFORE epoch seconds: the ranges do
    # not overlap, and getting the order wrong is the silent-1970 bug.
    if _looks_numeric(text):
        number = float(text)
        if EXCEL_SERIAL_MIN <= number <= EXCEL_SERIAL_MAX:
            return _from_excel_serial(number)
        # A fractional number outside the serial window is not a timestamp we know.
        # Refusing beats coercing it into a date that looks reasonable.
        if not float(number).is_integer():
            raise NormalizationError(
                f"{source_name} row {row_num}: {value!r} is fractional but outside the "
                f"Excel serial-date range ({EXCEL_SERIAL_MIN}–{EXCEL_SERIAL_MAX}). "
                "Refusing to guess whether it is a date."
            )

    # Epoch seconds — what Razorpay's API returns.
    if text.isdigit():
        return datetime.fromtimestamp(int(text), tz=UTC)

    # "%d/%m/%Y %H:%M:%S" and "%d-%m-%Y %H:%M:%S" are the shapes Razorpay's dashboard
    # exports use alongside the serials. Longest-first so a datetime is not truncated
    # to a bare date by an earlier partial match.
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue

    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        raise NormalizationError(
            f"{source_name} row {row_num}: cannot parse timestamp {value!r}. "
            "Accepted: Excel serial date, epoch seconds, YYYY-MM-DD, "
            "DD/MM/YYYY, or ISO 8601."
        ) from None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)


def normalize_ledger(
    path: Path, overrides: dict[str, str] | None = None
) -> tuple[list[dict[str, Any]], ColumnMapping]:
    """Read a merchant ledger CSV into canonical rows.

    Amounts arrive as rupee strings and leave as integer paise — this is the boundary.
    """
    return _normalize_csv(
        path,
        aliases=LEDGER_ALIASES,
        required=LEDGER_REQUIRED,
        source_name="ledger",
        money_fields={"amount_paise": False},   # False = negatives not allowed
        timestamp_fields=("captured_at",),
        overrides=overrides,
    )


def normalize_bank(
    path: Path, overrides: dict[str, str] | None = None
) -> tuple[list[dict[str, Any]], ColumnMapping]:
    """Read a bank statement CSV into canonical rows.

    Credits may legitimately be negative — a settlement reversal debits the account —
    so negatives are permitted here and refused in the ledger.
    """
    rows, mapping = _normalize_csv(
        path,
        aliases=BANK_ALIASES,
        required=BANK_REQUIRED,
        source_name="bank",
        money_fields={"credit_paise": True},
        timestamp_fields=(),
        overrides=overrides,
    )
    for i, row in enumerate(rows, start=2):
        raw = row.get("value_date")
        if raw:
            parsed = _parse_timestamp(raw, "bank", i)
            row["value_date"] = parsed.date() if parsed else None
    return rows, mapping


# File formats we can read. Razorpay's dashboard exports .xlsx; a merchant's own
# bookkeeping is usually .csv. Both must work, because "export your settlement report"
# hands a merchant an Excel file and an upload path that rejects it stops them on step
# one. See ADR-043.
TABULAR_SUFFIXES = frozenset({".csv", ".xlsx", ".xlsm"})


def _read_tabular(path: Path, source_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Read a CSV or Excel file into (headers, row dicts).

    Returns raw cell values. Every mapping, money and timestamp decision happens
    downstream, so the two formats cannot diverge in how a value is interpreted — the
    only difference this function is allowed to introduce is where the bytes came from.
    """
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        return _read_excel(path, source_name)

    if suffix and suffix not in TABULAR_SUFFIXES:
        raise NormalizationError(
            f"{source_name}: cannot read {path.name} — unsupported format {suffix!r}. "
            f"Accepted: {', '.join(sorted(TABULAR_SUFFIXES))}."
        )

    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = list(reader.fieldnames or [])
        return headers, list(reader)


def _read_excel(path: Path, source_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Read the first worksheet of an .xlsx/.xlsm file.

    `data_only=True` returns the cached value of a formula rather than the formula
    text. A settlement report with a SUM in it must yield the number, not "=SUM(A1:A9)".

    Cells are handed on as-is: openpyxl returns real `datetime` objects for
    date-formatted cells and floats for numbers, and `_parse_timestamp` and
    `parse_money` both accept those. Stringifying here would throw away type
    information and re-create the Excel-serial ambiguity of ADR-037.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:   # pragma: no cover - dependency is declared
        raise NormalizationError(
            f"{source_name}: reading {path.name} needs openpyxl. "
            "Install the engine's dependencies."
        ) from exc

    try:
        # Razorpay's exports carry no default style block, which openpyxl warns about on
        # every single file. The warning is about cosmetics we never read — we take cell
        # VALUES — so it is noise that would train a user to ignore warnings.
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=".*no default style.*")
            wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise NormalizationError(
            f"{source_name}: cannot open {path.name} as an Excel workbook: {exc}"
        ) from exc

    try:
        ws = wb.worksheets[0]
        grid = ws.iter_rows(values_only=True)

        try:
            header_row = next(grid)
        except StopIteration:
            return [], []

        # Razorpay's settlements export opens with a blank spacer column, so trailing
        # and empty headers are dropped rather than becoming a column named "None".
        headers = [
            str(h).strip() for h in header_row if h is not None and str(h).strip()
        ]
        width = len(headers)

        rows: list[dict[str, Any]] = []
        for values in grid:
            # A wholly empty row is spacing, not data. Excel files are full of them, and
            # one would otherwise become a row of empty strings that fails money parsing.
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            rows.append(dict(zip(headers, values[:width], strict=False)))

        return headers, rows
    finally:
        wb.close()


def _normalize_csv(
    path: Path,
    *,
    aliases: dict[str, tuple[str, ...]],
    required: tuple[str, ...],
    source_name: str,
    money_fields: dict[str, bool],
    timestamp_fields: tuple[str, ...],
    overrides: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], ColumnMapping]:
    if not path.exists():
        raise NormalizationError(f"{source_name}: file not found: {path}")

    headers, raw_rows = _read_tabular(path, source_name)
    if not headers:
        raise NormalizationError(
            f"{source_name}: {path} has no header row. Refusing to read positionally."
        )

    mapping = resolve_columns(headers, aliases, required, source_name, overrides)
    rows: list[dict[str, Any]] = []

    for row_num, raw in enumerate(raw_rows, start=2):
        out: dict[str, Any] = {}
        for canonical, input_col in mapping.resolved.items():
            value = raw.get(input_col)

            if canonical in money_fields:
                try:
                    out[canonical] = parse_money(
                        value if value not in (None, "") else 0,
                        allow_negative=money_fields[canonical],
                    )
                except MoneyError as exc:
                    raise NormalizationError(
                        f"{source_name} row {row_num}, column {input_col!r}: {exc}"
                    ) from exc
            elif canonical in timestamp_fields:
                out[canonical] = _parse_timestamp(value, source_name, row_num)
            else:
                out[canonical] = value.strip() if isinstance(value, str) else value

        out["_row"] = row_num
        rows.append(out)

    return rows, mapping


def load_collection(path: Path, source_name: str) -> list[dict[str, Any]]:
    """Read a Razorpay collection envelope: {"entity": "collection", "items": [...]}.

    Already canonical — these are Razorpay's own field names (ADR-008), which is the
    whole point of not renaming them. Accepts a bare list too, since a paginated fetch
    may be concatenated before it reaches us.
    """
    if not path.exists():
        raise NormalizationError(f"{source_name}: file not found: {path}")
    try:
        data = json.loads(path.read_text())
    except json.JSONDecodeError as exc:
        raise NormalizationError(f"{source_name}: invalid JSON in {path}: {exc}") from exc

    if isinstance(data, list):
        return data
    if not isinstance(data, dict) or "items" not in data:
        raise NormalizationError(
            f"{source_name}: expected a Razorpay collection with an 'items' key, "
            f"got keys {sorted(data) if isinstance(data, dict) else type(data).__name__}"
        )

    items = data["items"]
    declared = data.get("count")
    if declared is not None and declared != len(items):
        # A truncated page would silently under-report every total downstream.
        raise NormalizationError(
            f"{source_name}: collection declares count={declared} but carries "
            f"{len(items)} items. Refusing a partial page."
        )
    return items


def to_date(value: Any) -> date | None:
    """Coerce a timestamp-ish value to a UTC date. Used by the matcher and classifier.

    Delegates the string case to `_parse_timestamp`, which is the engine's real date
    parser — Excel serials, epoch seconds, ISO, DD/MM/YYYY and DD-MM-YYYY, with an error
    that names the accepted formats.

    This function used to do its own parsing and reached only `date.fromisoformat`, so
    it raised a bare `ValueError: Invalid isoformat string: '29/06/2022 07:34:39'` on a
    string taken verbatim from Razorpay's own settlement export — a format
    `_parse_timestamp` has read correctly since ADR-044, in the very column its docstring
    cites. Two date parsers, one of them good.

    Found by running the arithmetic tests against `sample-settlements-recon-report.xlsx`
    rather than against generated data: the generator writes one timestamp format per
    column because nobody would think to generate a column that mixes two. See ADR-056.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        # A NAIVE datetime is already UTC and is stamped as such, not converted.
        # `openpyxl` returns naive datetimes for every date cell in an .xlsx, and
        # `.astimezone(UTC)` interprets a naive value as LOCAL time — so on an IST
        # machine (+5:30) a settlement stamped 02:00 became the previous day, and the
        # engine reported a settlement a day earlier than the file says. Silent, machine-
        # dependent, and exactly the off-by-one on a settlement date that this engine
        # exists to catch in other people's systems.
        #
        # The rest of the engine treats Razorpay timestamps as UTC (`_parse_timestamp`
        # stamps `tzinfo=UTC` rather than converting), so this is consistency, not a
        # new assumption. ADR-056.
        if value.tzinfo is None:
            return value.date()
        return value.astimezone(UTC).date()
    if isinstance(value, date):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return datetime.fromtimestamp(value, tz=UTC).date()
    if isinstance(value, str):
        # `_parse_timestamp` wants a source and row for its message; this function is
        # called from the matcher and classifier, which have neither. The value itself
        # is what a reader needs to find the offending cell.
        parsed = _parse_timestamp(value, f"date value {value!r}", 0)
        return parsed.date() if parsed else None
    raise NormalizationError(f"cannot interpret {value!r} as a date")
